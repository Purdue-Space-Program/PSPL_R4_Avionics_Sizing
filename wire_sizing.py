import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook


####################################################################
# Wire Sizing Script
#
# Author: Jeremy Shinn
# Contact me on slack or via email Shinn3@purdue.edu
#
# Descritpion: This code sizes wires for given current requirements
# visit the wire sizing confluence page for more information
#
####################################################################

def CalcTemp(PreviousConTemp, PreviousInTemp):


    Resistance = (Resistivity * WireLength) * (1 + ConTempResistCoeff * (PreviousConTemp - AirTemp))
    
    PreviousConEnergy = ConMass * ConHeatCapacity * (PreviousConTemp - AirTemp)
    PreviousInEnergy = InMass * InHeatCapacity * (PreviousInTemp - AirTemp)
    ConEnergyIn = (Current ** 2) * Resistance * TimeStep
    
    InEnergyIn = TimeStep * InConductiveCoefficient * ConSurfaceArea * (PreviousConTemp - PreviousInTemp) / InThickness 
    InEnergyOut = AirConvectiveCoefficient * InSurfaceArea * (PreviousInTemp - AirTemp) * TimeStep
    
    InEnergyStored = InEnergyIn - InEnergyOut + PreviousInEnergy
    ConEnergyStored = ConEnergyIn - InEnergyIn + PreviousConEnergy
    ConTemp = (ConEnergyStored / (ConMass * ConHeatCapacity)) + AirTemp
    InTemp = (InEnergyStored / (InMass * InHeatCapacity)) + AirTemp
    return ConTemp, InTemp

###########################################################################
# Define parameters
# User input
WireLength = 3 # [M] Length of the wire
InitialTemp = 293.15 # [K] Initial temperature of the wire

#Spreadsheet values
wb = load_workbook('wire_sizing_parameters.xlsx')
ws = wb.active
Resistivity = ws['C10'].value # [ohm/Meter] Resistivity of the conductor
ConDensity = ws['C11'].value # [kg/m^3] Density of conductor
ConHeatCapacity = ws['C12'].value # [J/(kgK)] Specific heat capacity of conductor
ConTempResistCoeff = ws['C13'].value # [1/C] Temperature coefficient of resistance for conductor material
ConRadius = ws['C14'].value # [m] Radius of the wire conductor
InThickness = ws['C8'].value # [m] Outer radius minus inner radius
InConductiveCoefficient = ws['C6'].value # [W/mK] Thermal conductivity of the insulative material
InDensity = ws['C5'].value # [kg/m^3] Density of the insulative material
InHeatCapacity = ws['C7'].value # [J/kgK] Specific heat capacity of the insulative material

#Calculated
ConCrossArea = 3.14 * ConRadius ** 2 # [m^2] Cross sectional area of the condcutor
ConSurfaceArea = 2 * 3.14 *ConRadius * WireLength # [m^2] Total surface area of the conductor
ConMass = ConCrossArea * WireLength * ConDensity # [kg] Mass of the conductor
TimeStep = .1 # [Seconds]
AirTemp = 293 # [K] temperature of the air surrounding the wire

InCrossArea = 3.14 * (ConRadius + InThickness) ** 2 - ConCrossArea # [m^2] Cross-sectional area of the insulative material
InMass = InDensity * InCrossArea # [kg] Mass of the insulative material
InSurfaceArea = 2 * 3.14 * (ConRadius + InThickness) * WireLength # [m^2] Outer surface area of the insulator

InitialConTemp = InitialTemp # [K] Initial temperature of the conductor
InitialInTemp = InitialTemp # [K] Initial temperature of the insulation
AirConvectiveCoefficient = 9 # [W/(m^2C)] Convective heat transfer coefficient
PreviousConTemp = 0
TimeEnd = -1
ConTempArray = []
InTempArray = []
CountArray = []
############################################################################

#################################################################
# Get inputs
Current = int(input("Please enter current requirements [Amps]: "))
Choice = str(input("Would you like to enter an ending time? [Y/N]: "))
if Choice.lower() == "y":    
    TimeEnd = int(input("Please enter an ending time [Second]: "))

####################################################################

Energyin = []
EnergyOut = []

count = 0
PreviousTemps = CalcTemp(InitialConTemp, InitialInTemp)
while True:
    if count == TimeEnd * 10:
        break
    Temperatures = CalcTemp(PreviousTemps[0],PreviousTemps[1])
    
    ConTempArray.append(Temperatures[0] - 273.15)
    InTempArray.append(Temperatures[1] - 273.15)
    CountArray.append(count)
    if (Temperatures[0] - PreviousTemps[0]) < .0000001:
        print(f"\nSteady state conductor temperature: {str(round(Temperatures[0] - 273.15, 3))} C")
        print(f"\nSteady state insulation temperature: {str(round(Temperatures[1] - 273.15, 3))} C")      
        break
    if Temperatures[0] >  10000:
        print("Wire will melt before steady state")
        break
    PreviousTemps = Temperatures
    count += 1

TimeArray = list(range(0, int(count/10)))
ConTempArray = ConTempArray[9::10]
InTempArray = InTempArray[9::10]

plt.plot(TimeArray, ConTempArray, color = "red", label = 'Conductor')
plt.title("Wire Temperature", fontsize = 16, fontweight = "bold")
plt.xlabel("Time (Seconds)")
plt.ylabel("Temperature (Celsius)")
plt.grid(color='gray', linestyle='-', linewidth=0.5, alpha=0.7)
plt.plot(TimeArray, InTempArray, color = "blue", label = 'Insulation')
plt.legend()

plt.show()

